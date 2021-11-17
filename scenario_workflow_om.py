import os
import sys

from openpyxl import load_workbook
import pandas as pd

import get_present_settings
import get_template_settings
import scenario_from_csv


EXCEL = "Scenario workflow Over Morgen.xlsx"

def call():
    """
    Make sure the Excel is closed when calling this function
    """
    wb = load_workbook(EXCEL)
    df = pd.read_excel(EXCEL, sheet_name=None)

    get_area(wb, df)
    get_templates(wb, df)
    create_scenario(wb, df)


def get_area(wb, df):
    """
    Process the area in order to get the start values
    """

    # Copy-paste area into the areas.csv file in the data/input directory
    try:
        df['areas.csv'].to_csv('data/input/areas.csv', index=False)
        area = df['areas.csv']['geo_id'][0]
    except KeyError:
        print("De tab \"areas.csv\" kan niet worden gevonden in Excel.")

    # Get present settings
    get_present_settings.call()

    # Based on area name find the present settings file
    present_settings = pd.read_csv(f'data/output/present_settings_{area}.csv', index_col=0)

    # Copy-paste present settings into the Excel (sheet "Startwaarden")
    if 'Startinstellingen' in wb.sheetnames:
        wb.remove(wb['Startinstellingen'])
        print(f"\nOude startinstellingen zijn verwijderd.")

    writer = pd.ExcelWriter(EXCEL, engine='openpyxl', index=False)
    writer.book = wb
    present_settings.to_excel(writer, sheet_name='Startinstellingen')
    writer.save()
    writer.close()

    print("\nStartinstellingen zijn geüpdatet in Excel!")


def get_templates(wb, df):
    """
    Process the templates
    """

    for year in ['2030', '2040', '2050']:

        try:
            # Copy-paste template_settings.csv file into the data/input directory
            df[f'template_{year}.csv'].to_csv('data/input/template_list.csv', index=False)
            template_title = df[f'template_{year}.csv']['title'][0]
            print(f"\nTemplate voor {year} is ingelezen.")

            # Get template settings
            get_template_settings.call()

            # Find template settings file
            template_settings = pd.read_csv(f'data/output/template_settings.csv', index_col=0)

            # Copy-paste present settings into the Excel (sheetname equal to template title)
            if template_title in wb.sheetnames:
                wb.remove(wb[template_title])
                print(f"\nOude template instellingen voor {template_title} zijn verwijderd.")

            writer = pd.ExcelWriter(EXCEL, engine='openpyxl', index=False)
            writer.book = wb
            template_settings.to_excel(writer, sheet_name=template_title)
            writer.save()
            writer.close()

            print(f"\nTemplate instellingen voor {year} zijn geüpdatet in Excel!")

        except:
            print(f"\nWaarschuwing! De tab \"template_{year}.csv\" kan niet worden gevonden of geüpdatet in Excel.\n")
            pass


def create_scenario(wb, df):
    """
    Create scenario based on the scenario list and settings
    """

    # Copy-paste scenario_list.csv and scenario_settings.csv from Excel into
    # the data/input directory
    try:
        df['scenario_list.csv'].to_csv('data/input/scenario_list.csv', index=False)
    except KeyError:
        print("De tab \"scenario_list.csv\" kan niet worden gevonden in Excel.")

    try:
        df['scenario_settings.csv'].to_csv('data/input/scenario_settings.csv', index=False)
    except KeyError:
        print("De tab \"scenario_settings.csv\" kan niet worden gevonden in Excel.")

    # Create scenario
    scenario_from_csv.call()
    print("\nScenario's zijn aangemaakt!")

    # Copy-paste scenario_list.csv back into Excel
    scenario_list = pd.read_csv(f'data/input/scenario_list.csv', index_col=0)

    if 'scenario_list.csv' in wb.sheetnames:
        wb.remove(wb['scenario_list.csv'])

    writer = pd.ExcelWriter(EXCEL, engine='openpyxl', index=False)
    writer.book = wb
    scenario_list.to_excel(writer, sheet_name='scenario_list.csv')
    writer.save()
    writer.close()

    print("\nDe tab \"scenario_list.csv\" is geüpdatet in Excel!")


if __name__ == "__main__":
    call()
